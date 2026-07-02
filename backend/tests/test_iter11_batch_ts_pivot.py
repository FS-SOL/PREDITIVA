"""Iter11 — BUG-fix regression tests for:
(1) POST /api/measurements/import and POST /api/thermal/import must use a SINGLE batch timestamp
    (batch_ts) for ALL rows of a given import (fixes "escada" of unique per-row timestamps).
(2) GET /api/measurements/export and GET /api/thermal/export must group date columns by MINUTE
    (_fmt_dt) — no duplicate columns for the same minute even if the raw DB data has microsecond drift.
"""
import io
import os
import uuid
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL="):
                BASE_URL = line.split("=", 1)[1].strip().rstrip("/")

ADMIN = {"email": "admin@fssolucoes.com", "password": "admin123"}


@pytest.fixture(scope="module")
def admin_client():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login", json=ADMIN, timeout=15)
    assert r.status_code == 200, r.text
    s.headers.update({"Authorization": f"Bearer {r.json()['token']}"})
    return s


# ---------------------------------------------------------------- helpers ---
def _mk_vib_xlsx(equip: str, sub: str, rows: list) -> bytes:
    """Build a vibração import xlsx with rows [(ponto, valor, unidade, deteccao), ...]."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Equipamento", "Subconjunto", "Ponto", "Unidade", "Detecção", "Valor"])
    for (ponto, valor, unidade, deteccao) in rows:
        ws.append([equip, sub, ponto, unidade, deteccao, valor])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _mk_therm_xlsx(equip: str, rows: list) -> bytes:
    """Build a thermal import xlsx with rows [(ponto, temperatura, temp_ambiente), ...]."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Equipamento", "Ponto", "Temperatura", "Temp_Ambiente"])
    for (ponto, t, amb) in rows:
        ws.append([equip, ponto, t, amb])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============== A) Vibração import → single batch_ts =====================
class TestVibracaoBatchTs:
    def test_multi_points_share_single_timestamp(self, admin_client):
        # Create a temp machine (parent+sub) to isolate
        parent_tag = f"TEST_TS_{uuid.uuid4().hex[:6].upper()}"
        sub = "Motor"
        payload = {
            "tag": f"{parent_tag} / {sub}",
            "parent_tag": parent_tag,
            "subconjunto": sub,
            "local": "TEST",
            "equipamento": parent_tag,
            "descricao": sub,
            "tipo": "vibracao",
            "criticidade": "Média",
        }
        r = admin_client.post(f"{BASE_URL}/api/machines", json=payload, timeout=15)
        assert r.status_code in (200, 201), r.text
        machine = r.json()
        mid = machine["id"]

        try:
            # Import an xlsx with 8 different pontos → all must share same 'data'
            rows = [(f"P{i}H", 1.0 + i * 0.1, "mm/s", "Velocidade") for i in range(8)]
            content = _mk_vib_xlsx(parent_tag, sub, rows)
            files = {"file": ("import.xlsx", content,
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            r = admin_client.post(f"{BASE_URL}/api/measurements/import", files=files, timeout=30)
            assert r.status_code == 200, r.text
            body = r.json()
            assert body.get("inserted", 0) >= 8, body

            # Query measurements for this machine → collect distinct 'data' values
            r = admin_client.get(f"{BASE_URL}/api/measurements",
                                 params={"machine_id": mid}, timeout=15)
            assert r.status_code == 200
            meas = r.json()
            assert len(meas) >= 8, f"expected >=8 rows, got {len(meas)}"
            # Pick only rows from this machine that came from our import (all of them, as machine is new)
            distinct = set(m["data"] for m in meas)
            assert len(distinct) == 1, (
                f"BUG: escada — imported measurements have {len(distinct)} distinct timestamps, "
                f"expected 1. Values: {distinct}"
            )
        finally:
            # cleanup — delete the temp machine and its measurements
            admin_client.delete(f"{BASE_URL}/api/machines/{mid}", timeout=15)

    def test_two_imports_produce_two_distinct_timestamps(self, admin_client):
        """Two separate imports → two distinct batch_ts; ensures we don't collapse across imports."""
        import time as _t
        parent_tag = f"TEST_TS2_{uuid.uuid4().hex[:6].upper()}"
        sub = "Motor"
        payload = {
            "tag": f"{parent_tag} / {sub}",
            "parent_tag": parent_tag,
            "subconjunto": sub,
            "tipo": "vibracao",
            "criticidade": "Média",
        }
        r = admin_client.post(f"{BASE_URL}/api/machines", json=payload, timeout=15)
        assert r.status_code in (200, 201), r.text
        mid = r.json()["id"]
        try:
            for i in range(2):
                rows = [(f"P{j}H", float(j + 1), "mm/s", "Velocidade") for j in range(3)]
                content = _mk_vib_xlsx(parent_tag, sub, rows)
                files = {"file": ("import.xlsx", content,
                                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
                r = admin_client.post(f"{BASE_URL}/api/measurements/import", files=files, timeout=30)
                assert r.status_code == 200, r.text
                _t.sleep(1.1)  # ensure different second → different batch_ts
            r = admin_client.get(f"{BASE_URL}/api/measurements",
                                 params={"machine_id": mid}, timeout=15)
            distinct = set(m["data"] for m in r.json())
            assert len(distinct) == 2, (
                f"Two separate imports should yield 2 distinct timestamps; got {len(distinct)}"
            )
        finally:
            admin_client.delete(f"{BASE_URL}/api/machines/{mid}", timeout=15)


# ============== B) Thermal import → single batch_ts ======================
class TestThermalBatchTs:
    def test_multi_points_share_single_timestamp(self, admin_client):
        parent_tag = f"TEST_TH_{uuid.uuid4().hex[:6].upper()}"
        payload = {
            "tag": parent_tag,
            "parent_tag": parent_tag,
            "subconjunto": "",
            "tipo": "termografia",
            "criticidade": "Média",
        }
        r = admin_client.post(f"{BASE_URL}/api/machines", json=payload, timeout=15)
        assert r.status_code in (200, 201), r.text
        mid = r.json()["id"]
        try:
            rows = [(f"Fase {p}", 40.0 + i * 3, 25.0) for i, p in enumerate(["R", "S", "T", "N", "PE"])]
            content = _mk_therm_xlsx(parent_tag, rows)
            files = {"file": ("th.xlsx", content,
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            r = admin_client.post(f"{BASE_URL}/api/thermal/import", files=files, timeout=30)
            assert r.status_code == 200, r.text
            assert r.json().get("inserted", 0) >= 5

            r = admin_client.get(f"{BASE_URL}/api/thermal",
                                 params={"machine_id": mid}, timeout=15)
            items = r.json()
            assert len(items) >= 5
            distinct = set(x["data"] for x in items)
            assert len(distinct) == 1, (
                f"BUG: thermal import produced {len(distinct)} distinct timestamps; expected 1"
            )
        finally:
            admin_client.delete(f"{BASE_URL}/api/machines/{mid}", timeout=15)


# ============== C) Exports group columns by minute =======================
class TestExportPivotByMinute:
    def _load_export(self, admin_client, path):
        r = admin_client.get(f"{BASE_URL}{path}", timeout=30)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        return headers, ws

    def test_measurements_export_headers_no_duplicate_minute(self, admin_client):
        headers, ws = self._load_export(admin_client, "/api/measurements/export")
        assert headers[:4] == ["Equipamento", "Ponto", "Detecção", "Unidade"]
        date_cols = headers[4:]
        # No duplicates
        assert len(date_cols) == len(set(date_cols)), (
            f"Duplicate date columns in measurements export: {date_cols}"
        )

    def test_measurements_export_bba_single_column_per_minute(self, admin_client):
        """BBA-0001 / Motor (seed with microsecond-drift timestamps) must pivot to one
        column per unique minute (not per raw timestamp)."""
        # First figure machine_id
        r = admin_client.get(f"{BASE_URL}/api/machines", timeout=15)
        machines = r.json()
        bba = next((m for m in machines if m.get("tag") == "BBA-0001 / Motor"), None)
        if not bba:
            pytest.skip("BBA-0001 / Motor seed not present")
        headers, _ = self._load_export(admin_client, f"/api/measurements/export?machine_id={bba['id']}")
        date_cols = headers[4:]
        # Query measurements → compute distinct minute keys
        r = admin_client.get(f"{BASE_URL}/api/measurements",
                             params={"machine_id": bba["id"]}, timeout=15)
        meas = r.json()
        from datetime import datetime as _dt

        def _minute(iso):
            try:
                return _dt.fromisoformat((iso or "").replace("Z", "+00:00")).strftime("%d/%m/%Y %H:%M")
            except Exception:
                return iso

        distinct_minutes = set(_minute(m["data"]) for m in meas)
        assert len(date_cols) == len(distinct_minutes), (
            f"Export column count {len(date_cols)} != distinct minutes {len(distinct_minutes)}. "
            f"cols={date_cols} minutes={distinct_minutes}"
        )
        # No duplicate columns
        assert len(date_cols) == len(set(date_cols))

    def test_thermal_export_headers_no_duplicate_minute(self, admin_client):
        headers, ws = self._load_export(admin_client, "/api/thermal/export")
        assert headers[:4] == ["Equipamento", "Ponto", "Detecção", "Unidade"]
        date_cols = headers[4:]
        assert len(date_cols) == len(set(date_cols)), (
            f"Duplicate date columns in thermal export: {date_cols}"
        )

    def test_thermal_export_qde_three_columns(self, admin_client):
        """QDE-229.01 seed → thermal timeline has 3 distinct minutes (10/04, 10/05, 10/06)."""
        r = admin_client.get(f"{BASE_URL}/api/machines", timeout=15)
        qde = next((m for m in r.json() if m.get("tag") == "QDE-229.01"), None)
        if not qde:
            pytest.skip("QDE-229.01 seed not present")
        headers, _ = self._load_export(admin_client, f"/api/thermal/export?machine_id={qde['id']}")
        date_cols = headers[4:]
        assert len(date_cols) == 3, f"Expected 3 date columns for QDE-229.01, got {date_cols}"

    def test_measurements_export_esf_four_columns(self, admin_client):
        """ESF-0001 / Motor 01 seed → 4 distinct minutes (15/04, 15/05, 15/06 09:00, 15/06 15:30
        or similar). Ensures different minutes on the SAME day are NOT collapsed."""
        r = admin_client.get(f"{BASE_URL}/api/machines", timeout=15)
        esf = next((m for m in r.json() if m.get("tag") == "ESF-0001 / Motor 01"), None)
        if not esf:
            pytest.skip("ESF-0001 / Motor 01 seed not present")
        headers, _ = self._load_export(admin_client, f"/api/measurements/export?machine_id={esf['id']}")
        date_cols = headers[4:]
        assert len(date_cols) == 4, f"Expected 4 date columns for ESF-0001/Motor 01, got {date_cols}"
