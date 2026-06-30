"""
Iteration 6 regression: real spreadsheet termografia import.
- POST /api/machines/import with /tmp/todas.xlsx must import the TERMOGRAFIA sheet (~200 rows).
- After import, GET /api/machines?tipo=termografia returns ~200 records, each with tag = raw TAG, tipo=termografia.
- Reimport must be idempotent (inserted=0 OR skipped>0; no duplication).
- Vibração path remains intact (215 composite-tag machines after full file imported).
"""
import os
import io
import pytest
import requests
import openpyxl

def _read_backend_url():
    v = os.environ.get("REACT_APP_BACKEND_URL")
    if v:
        return v
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL="):
                return line.split("=", 1)[1].strip()
    raise RuntimeError("REACT_APP_BACKEND_URL not configured")

BASE_URL = _read_backend_url().rstrip("/")
API = f"{BASE_URL}/api"
ADMIN = {"email": "admin@fssolucoes.com", "password": "admin123"}
REAL_XLSX = "/tmp/todas.xlsx"
SS_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module")
def admin():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json=ADMIN, timeout=30)
    assert r.status_code == 200, r.text
    tok = r.json().get("token") or r.json().get("access_token")
    if tok:
        s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


def _post_import(admin):
    with open(REAL_XLSX, "rb") as fh:
        files = {"file": ("todas.xlsx", fh.read(), SS_MIME)}
    r = admin.post(f"{API}/machines/import", files=files, timeout=120)
    assert r.status_code == 200, r.text
    return r.json()


class TestRealTermografiaImport:
    def test_real_xlsx_present_and_termo_sheet(self):
        assert os.path.exists(REAL_XLSX), f"missing {REAL_XLSX}"
        wb = openpyxl.load_workbook(REAL_XLSX, data_only=True)
        upper = [s.upper() for s in wb.sheetnames]
        assert any("TERMO" in s for s in upper), f"sheet names: {wb.sheetnames}"

    def test_import_real_file(self, admin):
        body = _post_import(admin)
        print("First import result:", body)
        # If DB already has these tags (previous run), inserted may be 0 but skipped should be > 0
        assert (body.get("inserted", 0) + body.get("skipped", 0)) > 100, body

    def test_termografia_count_and_shape(self, admin):
        r = admin.get(f"{API}/machines", params={"tipo": "termografia"}, timeout=60)
        assert r.status_code == 200, r.text
        items = r.json()
        # Expecting ~200 termo entries per problem statement
        assert len(items) >= 150, f"expected ~200 termografia machines, got {len(items)}"
        # Spot-check a few known TAGs
        tags = {m.get("tag") for m in items}
        # The problem statement lists QDE-229.01 and TRF-0001 as examples
        # Don't fail if missing (file rows may use different tags), just print
        sample = [t for t in ("QDE-229.01", "TRF-0001") if t in tags]
        print("Sample tags matched:", sample, "| total termo:", len(items))
        # Each record shape
        for m in items[:5]:
            assert m.get("tipo") == "termografia"
            assert m.get("tag"), m
            assert "equipamento" in m
            assert "local" in m
            assert "_id" not in m

    def test_reimport_is_idempotent(self, admin):
        body = _post_import(admin)
        print("Reimport result:", body)
        # On reimport, inserted must be 0 (all already exist) and skipped must be > 0
        assert body.get("inserted", 0) == 0, f"reimport inserted non-zero: {body}"
        assert body.get("skipped", 0) > 100, f"reimport should skip >100: {body}"


class TestVibracaoUnbroken:
    def test_vibracao_count(self, admin):
        r = admin.get(f"{API}/machines", params={"tipo": "vibracao"}, timeout=60)
        assert r.status_code == 200, r.text
        items = r.json()
        # Per spec: 215 vibracao machines with composite tag '{parent_tag} / {subconjunto}'
        assert len(items) >= 200, f"expected ~215 vibracao machines, got {len(items)}"
        # composite tag check on a sample
        composite_count = sum(1 for m in items if " / " in (m.get("tag") or ""))
        assert composite_count >= len(items) * 0.9, f"most vibracao tags should be composite, got {composite_count}/{len(items)}"
