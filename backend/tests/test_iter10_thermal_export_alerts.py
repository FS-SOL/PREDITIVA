"""Iter10 backend tests — Termografia Temperatura module, Export XLSX (Vibração/Termografia),
Dashboard latest_alerts block."""
import io
import os
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL").rstrip("/") if os.environ.get("REACT_APP_BACKEND_URL") else None

if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL="):
                BASE_URL = line.split("=", 1)[1].strip().rstrip("/")

ADMIN = {"email": "admin@fssolucoes.com", "password": "admin123"}


@pytest.fixture(scope="session")
def admin_client():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login", json=ADMIN, timeout=15)
    assert r.status_code == 200, r.text
    s.headers.update({"Authorization": f"Bearer {r.json()['token']}"})
    return s


# ============== A) Thermal template ==============
class TestThermalTemplate:
    def test_template_returns_xlsx(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/thermal/template", timeout=20)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert headers == ["Equipamento", "Ponto", "Temperatura", "Temp_Ambiente"], headers
        # Ensure at least 1 thermal machine pre-filled
        assert ws.max_row >= 2, "template deve conter máquinas de termografia pré-preenchidas"


# ============== B) Thermal list (alarme, delta_t) ==============
class TestThermalList:
    def test_thermal_alarme_and_delta_t(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/thermal", timeout=15)
        assert r.status_code == 200
        items = r.json()
        assert isinstance(items, list) and len(items) >= 6
        for it in items:
            assert "alarme" in it
            assert "delta_t" in it
        # QDE-229.01 Fase R latest = 95 → A2, ΔT=67
        fase_r_95 = [x for x in items if x.get("machine_tag") == "QDE-229.01" and x.get("ponto") == "Fase R" and x.get("temperatura") == 95]
        assert len(fase_r_95) == 1
        assert fase_r_95[0]["alarme"] == "A2"
        assert fase_r_95[0]["delta_t"] == 67.0
        # Fase S 72 → A1
        fase_s_72 = [x for x in items if x.get("machine_tag") == "QDE-229.01" and x.get("ponto") == "Fase S" and x.get("temperatura") == 72]
        assert len(fase_s_72) == 1
        assert fase_s_72[0]["alarme"] == "A1"


# ============== C) Thermal import (temp_alarm boundaries) ==============
class TestThermalImport:
    def _make_xlsx(self, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Equipamento", "Ponto", "Temperatura", "Temp_Ambiente"])
        for r in rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_import_and_classify(self, admin_client):
        # Use a temporary equipment so we can clean up (won't touch QDE-229.01 seed)
        tag = "TEST_TEMP_IMPORT_01"
        buf = self._make_xlsx([
            [tag, "P1", 45, 25],  # OK
            [tag, "P2", 75, 25],  # A1
            [tag, "P3", 105, 25], # A2
        ])
        files = {"file": ("temp.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = admin_client.post(f"{BASE_URL}/api/thermal/import", files=files, timeout=30)
        assert r.status_code == 200, r.text
        js = r.json()
        assert js["inserted"] == 3, js

        # Verify classification
        r2 = admin_client.get(f"{BASE_URL}/api/thermal", timeout=15)
        items = [x for x in r2.json() if x.get("machine_tag") == tag]
        by_ponto = {x["ponto"]: x for x in items}
        assert by_ponto["P1"]["alarme"] == "OK"
        assert by_ponto["P2"]["alarme"] == "A1"
        assert by_ponto["P3"]["alarme"] == "A2"
        assert by_ponto["P3"]["delta_t"] == 80.0

        # CLEANUP — delete thermal recs and the created stub machine
        for it in items:
            dr = admin_client.delete(f"{BASE_URL}/api/thermal/{it['id']}", timeout=15)
            assert dr.status_code == 200
        # remove the stub machine to keep DB tidy
        machines = admin_client.get(f"{BASE_URL}/api/machines?q={tag}", timeout=15).json()
        for m in machines:
            if m.get("tag") == tag:
                admin_client.delete(f"{BASE_URL}/api/machines/{m['id']}", timeout=15)

    def test_delete_thermal_by_id_and_point_clear_and_viewer_forbidden(self, admin_client):
        # create one thermal record via import, then hit DELETE by id and point/clear
        tag = "TEST_TEMP_DEL_01"
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["Equipamento", "Ponto", "Temperatura", "Temp_Ambiente"])
        ws.append([tag, "PA", 65, 25])
        ws.append([tag, "PA", 70, 25])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        files = {"file": ("t.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = admin_client.post(f"{BASE_URL}/api/thermal/import", files=files, timeout=30)
        assert r.status_code == 200

        items = [x for x in admin_client.get(f"{BASE_URL}/api/thermal").json() if x.get("machine_tag") == tag]
        assert len(items) == 2
        first_id = items[0]["id"]
        machine_id = items[0]["machine_id"]
        # DELETE by id
        dr = admin_client.delete(f"{BASE_URL}/api/thermal/{first_id}", timeout=15)
        assert dr.status_code == 200
        # point/clear the rest
        cr = admin_client.delete(f"{BASE_URL}/api/thermal/point/clear",
                                 params={"machine_id": machine_id, "ponto": "PA"}, timeout=15)
        assert cr.status_code == 200
        # verify empty
        remaining = [x for x in admin_client.get(f"{BASE_URL}/api/thermal").json() if x.get("machine_tag") == tag]
        assert remaining == []

        # cleanup stub machine
        for m in admin_client.get(f"{BASE_URL}/api/machines?q={tag}").json():
            if m.get("tag") == tag:
                admin_client.delete(f"{BASE_URL}/api/machines/{m['id']}", timeout=15)

        # Viewer must not be allowed. Create a viewer user via /auth/register
        s = requests.Session()
        r = s.post(f"{BASE_URL}/api/auth/register",
                   json={"email": "TEST_viewer_iter10@fs.com", "password": "viewer123",
                         "name": "Test Viewer", "role": "visualizador"}, timeout=15)
        # 400 if already exists — login instead
        if r.status_code != 200:
            r = s.post(f"{BASE_URL}/api/auth/login",
                       json={"email": "TEST_viewer_iter10@fs.com", "password": "viewer123"}, timeout=15)
            assert r.status_code == 200, r.text
        s.headers.update({"Authorization": f"Bearer {r.json()['token']}"})

        # DELETE by id should be 403
        # (any random id; require_editor rejects before DB lookup)
        rv = s.delete(f"{BASE_URL}/api/thermal/00000000-0000-0000-0000-000000000000", timeout=15)
        assert rv.status_code == 403, rv.text
        rv2 = s.delete(f"{BASE_URL}/api/thermal/point/clear",
                       params={"machine_id": "x", "ponto": "y"}, timeout=15)
        assert rv2.status_code == 403


# ============== D) Excel exports (Vibração + Termografia) ==============
class TestExports:
    def test_measurements_export_xlsx(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/measurements/export", timeout=30)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        # First 4 columns fixed
        assert headers[:4] == ["Equipamento", "Ponto", "Detecção", "Unidade"], headers
        # Should have at least 1 date column
        assert len(headers) >= 5
        # Should have rows
        assert ws.max_row >= 2

    def test_thermal_export_xlsx(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/thermal/export", timeout=30)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert headers[:4] == ["Equipamento", "Ponto", "Detecção", "Unidade"], headers
        assert len(headers) >= 5
        # rows: should have Fase R and Fase S for QDE-229.01
        tags = [(row[0].value, row[1].value) for row in ws.iter_rows(min_row=2)]
        assert ("QDE-229.01", "Fase R") in tags
        assert ("QDE-229.01", "Fase S") in tags


# ============== E) Dashboard latest_alerts ==============
class TestDashboardAlerts:
    def test_latest_alerts_contains_vibration_and_thermal(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/dashboard", timeout=20)
        assert r.status_code == 200
        d = r.json()
        assert "latest_alerts" in d
        alerts = d["latest_alerts"]
        assert isinstance(alerts, list)
        # only A2/Parado allowed
        for a in alerts:
            assert a["alarme"] in ("A2", "Parado")
            for k in ("tipo", "machine_tag", "ponto", "valor", "unidade", "alarme", "data"):
                assert k in a

        # Must contain the two expected alerts
        vib = [a for a in alerts if a["tipo"] == "Vibração" and a["machine_tag"] == "ESF-0001 / Motor 01" and a["ponto"] == "1HV"]
        assert len(vib) == 1
        assert vib[0]["valor"] == 10.5 and vib[0]["alarme"] == "A2" and vib[0]["unidade"] == "mm/s"

        ter = [a for a in alerts if a["tipo"] == "Termografia" and a["machine_tag"] == "QDE-229.01" and a["ponto"] == "Fase R"]
        assert len(ter) == 1
        assert ter[0]["valor"] == 95 and ter[0]["alarme"] == "A2" and ter[0]["unidade"] == "°C"
