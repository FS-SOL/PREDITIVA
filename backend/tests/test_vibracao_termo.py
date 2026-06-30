"""
Backend tests for measurements (vibration) and machines import (thermography).
Covers:
- GET /api/measurements/template (xlsx download)
- POST /api/measurements/import (case-insensitive match by parent_tag+subconjunto)
- GET /api/measurements (ISO 10816-3 alarm classification)
- DELETE /api/measurements/{id} and /api/measurements/point/clear (RBAC: viewer 403)
- POST /api/machines/import with sheet 'TERMOGRAFIA'
"""
import io
import os
import re
import uuid
import pytest
import requests
import openpyxl

# Read REACT_APP_BACKEND_URL from frontend/.env (no default)
def _read_backend_url():
    v = os.environ.get("REACT_APP_BACKEND_URL")
    if v:
        return v
    env_path = "/app/frontend/.env"
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                if line.startswith("REACT_APP_BACKEND_URL="):
                    return line.split("=", 1)[1].strip()
    raise RuntimeError("REACT_APP_BACKEND_URL not configured")

BASE_URL = _read_backend_url().rstrip("/")
API = f"{BASE_URL}/api"

ADMIN = {"email": "admin@fssolucoes.com", "password": "admin123"}
VIEWER = {"email": "viewer@fs.com", "password": "viewer123"}

SS_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _login(creds):
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json=creds, timeout=30)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    tok = r.json().get("token") or r.json().get("access_token")
    if tok:
        s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


@pytest.fixture(scope="module")
def admin():
    return _login(ADMIN)


@pytest.fixture(scope="module")
def viewer():
    try:
        return _login(VIEWER)
    except AssertionError as e:
        pytest.skip(f"viewer login unavailable: {e}")


# ---------- Helpers ----------

def _build_xlsx(headers, rows, sheet_name="Sheet1"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------- Tests ----------

class TestTemplate:
    def test_template_download(self, admin):
        r = admin.get(f"{API}/measurements/template", timeout=30)
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "")
        assert "spreadsheet" in ct or SS_MIME in ct, ct
        # parse content
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 2, "template should be pre-populated with machines"
        header = [str(c or "").strip().lower() for c in rows[0]]
        for h in ["equipamento", "subconjunto", "ponto", "unidade", "detecção", "valor"]:
            assert h in header, f"missing column {h} in {header}"


class TestImportMeasurements:
    inserted_meas_ids = []
    machine_used = None

    def test_import_case_insensitive_match(self, admin):
        # Get a real vibracao machine
        r = admin.get(f"{API}/machines", params={"tipo": "vibracao"}, timeout=30)
        assert r.status_code == 200
        machines = r.json()
        # Pick one with parent_tag + subconjunto populated
        target = None
        for m in machines:
            if m.get("parent_tag") and m.get("subconjunto"):
                target = m
                break
        assert target, "no vibracao machine with parent_tag+subconjunto found"
        TestImportMeasurements.machine_used = target

        parent_lower = (target["parent_tag"] or "").lower()
        sub = target["subconjunto"]

        # Build xlsx with 3 rows for ISO alarm coverage (1.5 OK, 8.2 A2, 12.5 Parado)
        rows = [
            [parent_lower, sub, "1HV", "mm/s", "Velocidade", 1.5],
            [parent_lower, sub, "2HA", "mm/s", "Velocidade", 8.2],
            [parent_lower, sub, "3HA", "mm/s", "Velocidade", 12.5],
        ]
        buf = _build_xlsx(
            ["Equipamento", "Subconjunto", "Ponto", "Unidade", "Detecção", "Valor"],
            rows,
            sheet_name="Overall",
        )
        files = {"file": ("import_test.xlsx", buf.getvalue(), SS_MIME)}
        r = admin.post(f"{API}/measurements/import", files=files, timeout=60)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body.get("inserted", 0) >= 3, body
        assert body.get("skipped", 0) == 0, body

        # GET measurements for this machine - last 3 should be the test rows with correct alarms
        r2 = admin.get(f"{API}/measurements", params={"machine_id": target["id"]}, timeout=30)
        assert r2.status_code == 200
        meas = r2.json()
        assert len(meas) >= 3
        # collect last 3
        last3 = meas[-3:]
        # ensure machine_tag is the composite tag, not raw parent
        for m in last3:
            assert m.get("machine_tag") == target["tag"], f"machine_tag mismatch: {m.get('machine_tag')} vs {target['tag']}"
        # alarms by value
        alarms = {round(m["valor"], 2): m["alarme"] for m in last3}
        assert alarms.get(1.5) == "OK", alarms
        assert alarms.get(8.2) == "A2", alarms
        assert alarms.get(12.5) == "Parado", alarms
        # capture ids for cleanup
        TestImportMeasurements.inserted_meas_ids = [m["id"] for m in last3]


class TestDeleteMeasurements:
    def test_delete_single_as_viewer_403(self, admin, viewer):
        if not TestImportMeasurements.inserted_meas_ids:
            pytest.skip("no measurement created")
        mid = TestImportMeasurements.inserted_meas_ids[0]
        r = viewer.delete(f"{API}/measurements/{mid}", timeout=30)
        assert r.status_code == 403, r.text

    def test_delete_single_as_admin(self, admin):
        if not TestImportMeasurements.inserted_meas_ids:
            pytest.skip("no measurement created")
        mid = TestImportMeasurements.inserted_meas_ids[0]
        r = admin.delete(f"{API}/measurements/{mid}", timeout=30)
        assert r.status_code == 200, r.text
        assert r.json().get("ok") is True

    def test_point_clear_as_viewer_403(self, viewer):
        m = TestImportMeasurements.machine_used
        if not m:
            pytest.skip("machine not selected")
        r = viewer.delete(
            f"{API}/measurements/point/clear",
            params={"machine_id": m["id"], "ponto": "2HA", "deteccao": "Velocidade"},
            timeout=30,
        )
        assert r.status_code == 403, r.text

    def test_point_clear_as_admin(self, admin):
        m = TestImportMeasurements.machine_used
        if not m:
            pytest.skip("machine not selected")
        # clear remaining 2HA and 3HA
        r1 = admin.delete(
            f"{API}/measurements/point/clear",
            params={"machine_id": m["id"], "ponto": "2HA", "deteccao": "Velocidade"},
            timeout=30,
        )
        assert r1.status_code == 200, r1.text
        body = r1.json()
        assert body.get("ok") is True
        assert isinstance(body.get("deleted"), int)
        r2 = admin.delete(
            f"{API}/measurements/point/clear",
            params={"machine_id": m["id"], "ponto": "3HA", "deteccao": "Velocidade"},
            timeout=30,
        )
        assert r2.status_code == 200


class TestMachinesImportTermo:
    inserted_tags = []

    def test_import_termografia_sheet(self, admin):
        # Compose unique TAG to avoid collision
        suffix = uuid.uuid4().hex[:6].upper()
        parent_tag = f"TEST-TERMO-{suffix}"
        rows = [
            # parent row with TAG+LOCAL+EQUIPAMENTOS, no COMPONENTE (will be skipped per current rule)
            [parent_tag, "Sala Teste", "Painel Teste", ""],
            # component rows - blank TAG and LOCAL (forward-fill), but with COMPONENTE
            [None, None, None, "Disjuntor A"],
            [None, None, None, "Disjuntor B"],
        ]
        buf = _build_xlsx(
            ["TAG", "LOCAL DE INSTALAÇÃO", "EQUIPAMENTOS", "COMPONENTE"],
            rows,
            sheet_name="TERMOGRAFIA",
        )
        files = {"file": ("termo_test.xlsx", buf.getvalue(), SS_MIME)}
        r = admin.post(f"{API}/machines/import", files=files, timeout=60)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body.get("inserted", 0) >= 2, body

        # Verify GET /api/machines?tipo=termografia includes new entries
        r2 = admin.get(f"{API}/machines", params={"tipo": "termografia"}, timeout=30)
        assert r2.status_code == 200
        all_termo = r2.json()
        matching = [m for m in all_termo if m.get("parent_tag") == parent_tag]
        assert len(matching) >= 2, f"expected new termo entries with parent {parent_tag}"
        TestMachinesImportTermo.inserted_tags = [m["id"] for m in matching]

    def test_cleanup_termo(self, admin):
        # Delete created machines via API to keep DB clean (best-effort)
        for mid in TestMachinesImportTermo.inserted_tags:
            admin.delete(f"{API}/machines/{mid}", timeout=15)
