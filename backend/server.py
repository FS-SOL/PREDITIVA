from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import re
import uuid
import logging
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Dict, Any

import bcrypt
import jwt
import openpyxl
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Query
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============== DB ==============
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_ALGORITHM = "HS256"
JWT_SECRET = os.environ.get("JWT_SECRET", "dev-secret-change")

app = FastAPI(title="FS Soluções - Preditiva")
api = APIRouter(prefix="/api")

# ============== Helpers ==============

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False

def iso_alarm(valor: float, unidade: str, deteccao: str) -> str:
    """ISO 10816-3 alarm classification.
    Velocity RMS (mm/s): OK<=2.8, A1<=7.1, A2<=11, Parado>11.
    Acceleration / envelope based on a simple g threshold."""
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return "OK"
    u = (unidade or "").lower()
    d = (deteccao or "").lower()
    is_velocity = "mm/s" in u or "velocidade" in d or "veloc" in d
    is_accel = "g" == u.strip() or "m/s" in u or "acelera" in d or "aceler" in d
    if is_velocity:
        if v <= 2.8:
            return "OK"
        if v <= 7.1:
            return "A1"
        if v <= 11.0:
            return "A2"
        return "Parado"
    if is_accel:
        if v <= 2.0:
            return "OK"
        if v <= 5.0:
            return "A1"
        if v <= 10.0:
            return "A2"
        return "Parado"
    # default (unknown unit): use velocity scale
    if v <= 2.8:
        return "OK"
    if v <= 7.1:
        return "A1"
    if v <= 11.0:
        return "A2"
    return "Parado"

def temp_alarm(temp) -> str:
    """Classificação de temperatura (°C): OK<=60, A1(Atenção)<=90, A2(Crítico)<=120, Parado>120."""
    try:
        t = float(temp)
    except (TypeError, ValueError):
        return "OK"
    if t <= 60:
        return "OK"
    if t <= 90:
        return "A1"
    if t <= 120:
        return "A2"
    return "Parado"

def _fmt_dt(iso: str) -> str:
    try:
        dt = datetime.fromisoformat((iso or "").replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return iso or ""

def _pivot_xlsx(rows, cols, title):
    """rows: list de dicts {machine_tag, ponto, deteccao, unidade, ordem, values:{colKey:val}}; cols: lista ordenada de rótulos."""
    import openpyxl as _op
    wb = _op.Workbook(); ws = wb.active; ws.title = title[:30]
    header = ["Equipamento", "Ponto", "Detecção", "Unidade"] + list(cols)
    ws.append(header)
    for c in range(1, len(header) + 1):
        ws.cell(row=1, column=c).font = _op.styles.Font(bold=True)
    rows_sorted = sorted(rows, key=lambda r: ((r.get("machine_tag") or ""), r.get("ordem", 0)))
    for r in rows_sorted:
        ws.append([r.get("machine_tag", ""), r.get("ponto", ""), r.get("deteccao", ""), r.get("unidade", "")] + [r["values"].get(c, "") for c in cols])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def create_token(user_id: str, email: str, ttl_min: int = 60 * 24 * 7) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=ttl_min),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Não autenticado")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Usuário não encontrado")
    return user

def require_editor(user=Depends(get_current_user)) -> dict:
    if user.get("role") == "visualizador":
        raise HTTPException(status_code=403, detail="Permissão negada: usuário somente leitura")
    return user

def require_admin(user=Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Permissão negada: requer administrador")
    return user

async def log_deletion(user: dict, entity_type: str, description: str, details: Optional[dict] = None):
    await db.deletion_logs.insert_one({
        "id": str(uuid.uuid4()),
        "entity_type": entity_type,
        "description": description,
        "details": details or {},
        "user_id": user.get("id"),
        "user_name": user.get("name", ""),
        "user_email": user.get("email", ""),
        "data": now_iso(),
    })

# ============== Models ==============
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = "tecnico"  # admin, tecnico, gestor, visualizador

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserOut(BaseModel):
    id: str
    email: str
    name: str
    role: str

class PlantNode(BaseModel):
    id: Optional[str] = None
    name: str
    type: str  # empresa, unidade, area, equipamento, subconjunto, ponto
    parent_id: Optional[str] = None

class MachineModel(BaseModel):
    id: Optional[str] = None
    tag: str
    subconjunto: Optional[str] = ""
    local: Optional[str] = ""
    equipamento: Optional[str] = ""
    descricao: Optional[str] = ""
    fabricante: Optional[str] = ""
    rpm: Optional[float] = None
    potencia: Optional[float] = None
    rolamento_loa: Optional[str] = ""
    rolamento_la: Optional[str] = ""
    criticidade: Optional[str] = "Média"  # Alta, Média, Baixa
    status: Optional[str] = "Sem diag."  # OK, A1, A2, Parado, Sem diag.
    tipo: Optional[str] = "vibracao"  # vibracao, termografia, ambos
    componente: Optional[str] = ""

class DefectModel(BaseModel):
    id: Optional[str] = None
    nome: str
    categoria: Optional[str] = ""  # mecanico, eletrico, rolamento, lubrificacao
    sintomas: Optional[str] = ""
    frequencias: Optional[str] = ""
    causas: Optional[str] = ""
    consequencias: Optional[str] = ""
    acoes: Optional[str] = ""
    diagnostico: Optional[str] = ""
    recomendacao: Optional[str] = ""
    alarme: Optional[str] = "A1"  # OK, A1, A2, Parado
    tipo: Optional[str] = "vibracao"  # vibracao, termografia, ambos
    ativo: bool = True

class DiagnosticoModel(BaseModel):
    id: Optional[str] = None
    machine_id: str
    machine_tag: Optional[str] = ""
    defect_ids: List[str] = []
    diagnostico: str = ""
    causa: str = ""
    consequencia: str = ""
    recomendacao: str = ""
    status: str = "A1"  # define new machine status
    tecnico: Optional[str] = ""
    data: Optional[str] = None

class MeasurementModel(BaseModel):
    id: Optional[str] = None
    machine_id: str
    machine_tag: Optional[str] = ""
    subconjunto: str
    ponto: str
    valor: float
    unidade: str
    deteccao: str
    ordem: Optional[int] = 0
    data: Optional[str] = None

# ============== Seeds ==============
DEFAULT_DEFECTS = [
    {"nome": "Desbalanceamento", "categoria": "mecanico", "sintomas": "Vibração predominante em 1xRPM, principalmente radial", "frequencias": "1xRPM (radial)", "causas": "Massa desigual no rotor, depósitos, parafusos perdidos", "consequencias": "Fadiga em rolamentos, trincas em estruturas, perda de eficiência", "acoes": "Balanceamento em campo, inspeção de rotor", "diagnostico": "Desbalanceamento de rotor confirmado por componente em 1xRPM", "recomendacao": "Programar balanceamento dinâmico no próximo PM", "alarme": "A1"},
    {"nome": "Desalinhamento", "categoria": "mecanico", "sintomas": "Vibração em 2xRPM axial elevada, vibração radial assimétrica", "frequencias": "1x, 2x e 3xRPM (axial e radial)", "causas": "Acoplamento mal alinhado, base deformada, dilatação térmica", "consequencias": "Desgaste prematuro de acoplamento e rolamentos", "acoes": "Realinhamento a laser", "diagnostico": "Desalinhamento entre motor e equipamento acionado", "recomendacao": "Executar alinhamento a laser e reavaliar em 30 dias", "alarme": "A1"},
    {"nome": "Rolamento - BPFO (Pista Externa)", "categoria": "rolamento", "sintomas": "Picos na frequência BPFO no envelope, modulações", "frequencias": "BPFO e harmônicas", "causas": "Falha por fadiga, contaminação, lubrificação inadequada", "consequencias": "Falha catastrófica do rolamento", "acoes": "Substituição do rolamento, revisar lubrificação", "diagnostico": "Defeito em pista externa do rolamento (BPFO)", "recomendacao": "Programar troca de rolamento dentro de 30-60 dias", "alarme": "A2"},
    {"nome": "Rolamento - BPFI (Pista Interna)", "categoria": "rolamento", "sintomas": "Picos em BPFI no envelope, bandas laterais em 1xRPM", "frequencias": "BPFI ± 1xRPM", "causas": "Fadiga, montagem inadequada, sobrecarga", "consequencias": "Falha do rolamento em curto prazo", "acoes": "Substituição imediata do rolamento", "diagnostico": "Defeito em pista interna do rolamento (BPFI)", "recomendacao": "Substituir rolamento na próxima parada programada", "alarme": "A2"},
    {"nome": "Rolamento - BSF (Esfera/Rolo)", "categoria": "rolamento", "sintomas": "Picos em BSF, modulação por FTF", "frequencias": "BSF ± FTF", "causas": "Defeito em elementos rolantes", "consequencias": "Ruído crescente, falha a curto/médio prazo", "acoes": "Substituição do rolamento", "diagnostico": "Defeito em elemento rolante (BSF)", "recomendacao": "Substituir rolamento", "alarme": "A2"},
    {"nome": "Folga Mecânica", "categoria": "mecanico", "sintomas": "Múltiplas harmônicas de 1xRPM (0.5x, 1x, 1.5x, 2x...)", "frequencias": "Harmônicas e sub-harmônicas de 1xRPM", "causas": "Fixação solta, desgaste em mancais", "consequencias": "Aumento de vibração geral, falhas em série", "acoes": "Verificar fixação, parafusos e mancais", "diagnostico": "Folga mecânica em base/mancal", "recomendacao": "Inspecionar e reapertar fixações", "alarme": "A1"},
    {"nome": "Cavitação", "categoria": "hidraulico", "sintomas": "Vibração aleatória em alta frequência, ruído característico", "frequencias": "Banda larga em alta frequência", "causas": "NPSH insuficiente, restrição na sucção", "consequencias": "Erosão de impulsor, perda de performance", "acoes": "Revisar sucção, NPSH, filtros", "diagnostico": "Cavitação em bomba", "recomendacao": "Verificar condições de sucção da bomba", "alarme": "A1"},
    {"nome": "Barra Quebrada (Motor)", "categoria": "eletrico", "sintomas": "Bandas laterais em ±2sLs em torno de 1xRPM", "frequencias": "1xRPM ± 2sLs", "causas": "Fadiga em barras do rotor, alta carga", "consequencias": "Perda de torque, falha do motor", "acoes": "Análise elétrica, rebobinamento", "diagnostico": "Barras quebradas no rotor do motor", "recomendacao": "Programar revisão elétrica do motor", "alarme": "A2"},
    {"nome": "Engrenagem - GMF", "categoria": "mecanico", "sintomas": "Picos em GMF e harmônicas, bandas laterais", "frequencias": "GMF ± Frequência do eixo", "causas": "Desgaste de dentes, dentes quebrados", "consequencias": "Falha de redutor", "acoes": "Inspecionar redutor", "diagnostico": "Defeito em engrenamento", "recomendacao": "Programar inspeção do redutor", "alarme": "A1"},
    {"nome": "Lubrificação Inadequada", "categoria": "lubrificacao", "sintomas": "Aumento gradual em alta frequência, envelope elevado", "frequencias": "Banda larga em alta frequência", "causas": "Falta ou excesso de graxa, contaminação", "consequencias": "Aceleração de falha em rolamentos", "acoes": "Relubrificar conforme plano", "diagnostico": "Lubrificação deficiente", "recomendacao": "Aplicar plano de lubrificação", "alarme": "A1"},
    {"nome": "Aquecimento Excessivo (Termografia)", "categoria": "termografia", "sintomas": "Temperatura acima do esperado em mancal/conexão", "frequencias": "—", "causas": "Sobrecarga, mau contato, lubrificação", "consequencias": "Falha do componente", "acoes": "Inspeção termográfica, ajustes", "diagnostico": "Ponto quente identificado", "recomendacao": "Inspecionar conexão/mancal", "alarme": "A1", "tipo": "termografia"},
    {"nome": "Equipamento em Bom Estado", "categoria": "geral", "sintomas": "Sem anomalias; níveis dentro dos limites normais", "frequencias": "—", "causas": "—", "consequencias": "—", "acoes": "Manter monitoramento periódico", "diagnostico": "Equipamento operando em condições normais. Nenhuma anomalia detectada nas medições/inspeção.", "recomendacao": "Manter plano de manutenção preditiva e a periodicidade de coleta.", "alarme": "OK", "tipo": "ambos"},
    {"nome": "Ponto Quente em Conexão Elétrica", "categoria": "termografia", "sintomas": "Sobreaquecimento localizado em terminal/borne comparado às demais fases", "frequencias": "—", "causas": "Aperto inadequado, oxidação, subdimensionamento", "consequencias": "Rompimento do condutor, incêndio, parada não planejada", "acoes": "Reapertar conexão, tratar oxidação, reavaliar", "diagnostico": "Ponto quente em conexão elétrica (ΔT elevado)", "recomendacao": "Programar correção da conexão na próxima parada", "alarme": "A2", "tipo": "termografia"},
    {"nome": "Sobrecarga / Sobreaquecimento", "categoria": "termografia", "sintomas": "Aquecimento generalizado do componente/cabo", "frequencias": "—", "causas": "Corrente acima da nominal, ventilação deficiente", "consequencias": "Degradação de isolação, redução de vida útil", "acoes": "Avaliar carga, balanceamento e ventilação", "diagnostico": "Sobrecarga térmica identificada", "recomendacao": "Revisar carregamento e dimensionamento", "alarme": "A1", "tipo": "termografia"},
    {"nome": "Mau Contato / Conexão Frouxa", "categoria": "termografia", "sintomas": "Aquecimento pontual em conexão, variável com a carga", "frequencias": "—", "causas": "Conexão frouxa, contato deficiente", "consequencias": "Falha do circuito, arco elétrico", "acoes": "Reaperto e limpeza da conexão", "diagnostico": "Mau contato em conexão", "recomendacao": "Corrigir conexão com equipe elétrica", "alarme": "A2", "tipo": "termografia"},
    {"nome": "Desequilíbrio de Fases", "categoria": "termografia", "sintomas": "Temperaturas diferentes entre as três fases", "frequencias": "—", "causas": "Carga desequilibrada, conexão deficiente em uma fase", "consequencias": "Aquecimento e perdas, atuação de proteção", "acoes": "Avaliar corrente por fase e equilibrar cargas", "diagnostico": "Desequilíbrio térmico entre fases", "recomendacao": "Balancear cargas do quadro", "alarme": "A1", "tipo": "termografia"},
]

async def seed_data():
    # Indexes
    await db.users.create_index("email", unique=True)
    await db.defects.create_index("nome", unique=True)
    await db.plant_nodes.create_index([("type", 1), ("parent_id", 1)])
    # Drop legacy unique index on tag (we now allow composite tag with subconjunto)
    try:
        await db.machines.drop_index("tag_1")
    except Exception:
        pass
    await db.machines.create_index([("tag", 1), ("subconjunto", 1)], unique=True)

    # Admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@fssolucoes.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "Administrador",
            "role": "admin",
            "created_at": now_iso(),
        })
    elif not verify_password(admin_password, existing.get("password_hash", "")):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})

    # Defects
    for d in DEFAULT_DEFECTS:
        tipo = d.get("tipo") or ("termografia" if d.get("categoria") == "termografia" else "vibracao")
        base = {k: v for k, v in d.items() if k != "tipo"}
        await db.defects.update_one(
            {"nome": d["nome"]},
            {"$set": {"tipo": tipo},
             "$setOnInsert": {**base, "id": str(uuid.uuid4()), "ativo": True, "created_at": now_iso()}},
            upsert=True,
        )
    # backfill tipo for any legacy defect missing it
    await db.defects.update_many({"tipo": {"$exists": False}}, {"$set": {"tipo": "vibracao"}})
    # backfill ordem for any legacy measurement missing it
    await db.measurements.update_many({"ordem": {"$exists": False}}, {"$set": {"ordem": 0}})

    # Empresa
    empresa = await db.plant_nodes.find_one({"type": "empresa"})
    if not empresa:
        await db.plant_nodes.insert_one({
            "id": str(uuid.uuid4()),
            "name": "FS Soluções",
            "type": "empresa",
            "parent_id": None,
            "created_at": now_iso(),
        })


@app.on_event("startup")
async def startup():
    await seed_data()
    logger.info("FS Soluções - Preditiva inicializado")

# ============== Auth Routes ==============
@api.post("/auth/register")
async def register(payload: UserCreate, response: Response):
    email = payload.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    user = {
        "id": str(uuid.uuid4()),
        "email": email,
        "password_hash": hash_password(payload.password),
        "name": payload.name,
        "role": payload.role,
        "created_at": now_iso(),
    }
    await db.users.insert_one(user)
    token = create_token(user["id"], email)
    response.set_cookie("access_token", token, httponly=True, secure=False, samesite="lax", max_age=60 * 60 * 24 * 7, path="/")
    return {"id": user["id"], "email": email, "name": user["name"], "role": user["role"], "token": token}

@api.post("/auth/login")
async def login(payload: UserLogin, response: Response):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    token = create_token(user["id"], email)
    response.set_cookie("access_token", token, httponly=True, secure=False, samesite="lax", max_age=60 * 60 * 24 * 7, path="/")
    return {"id": user["id"], "email": email, "name": user["name"], "role": user["role"], "token": token}

@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user

@api.get("/users")
async def list_users(user=Depends(get_current_user)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
    return users

# ============== Plant Hierarchy ==============
@api.get("/plants")
async def list_plants(user=Depends(get_current_user)):
    nodes = await db.plant_nodes.find({}, {"_id": 0}).to_list(5000)
    return nodes

@api.post("/plants")
async def create_plant(payload: PlantNode, user=Depends(require_editor)):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = now_iso()
    await db.plant_nodes.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/plants/{node_id}")
async def update_plant(node_id: str, payload: PlantNode, user=Depends(require_editor)):
    update = {k: v for k, v in payload.model_dump().items() if v is not None and k != "id"}
    res = await db.plant_nodes.update_one({"id": node_id}, {"$set": update})
    if not res.matched_count:
        raise HTTPException(404, "Nó não encontrado")
    return {"ok": True}

@api.delete("/plants/{node_id}")
async def delete_plant(node_id: str, user=Depends(require_editor)):
    node = await db.plant_nodes.find_one({"id": node_id}, {"_id": 0})
    await db.plant_nodes.delete_one({"id": node_id})
    if node:
        await log_deletion(user, "planta", f"{node.get('type','')}: {node.get('name','')}", {"id": node_id})
    return {"ok": True}

# ============== Machines ==============
@api.get("/machines")
async def list_machines(q: Optional[str] = None, tipo: Optional[str] = None, status: Optional[str] = None, user=Depends(get_current_user)):
    qry: Dict[str, Any] = {}
    if q:
        qry["$or"] = [
            {"tag": {"$regex": q, "$options": "i"}},
            {"equipamento": {"$regex": q, "$options": "i"}},
            {"local": {"$regex": q, "$options": "i"}},
            {"descricao": {"$regex": q, "$options": "i"}},
        ]
    if tipo and tipo != "todos":
        qry["tipo"] = {"$in": [tipo, "ambos"]}
    if status:
        qry["status"] = status
    items = await db.machines.find(qry, {"_id": 0}).to_list(5000)
    return items

@api.post("/machines")
async def create_machine(payload: MachineModel, user=Depends(require_editor)):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = now_iso()
    if await db.machines.find_one({"tag": doc["tag"]}):
        raise HTTPException(400, "TAG já cadastrada")
    await db.machines.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/machines/{machine_id}")
async def update_machine(machine_id: str, payload: MachineModel, user=Depends(require_editor)):
    update = {k: v for k, v in payload.model_dump().items() if k != "id"}
    res = await db.machines.update_one({"id": machine_id}, {"$set": update})
    if not res.matched_count:
        raise HTTPException(404, "Máquina não encontrada")
    return {"ok": True}

@api.delete("/machines/{machine_id}")
async def delete_machine(machine_id: str, user=Depends(require_editor)):
    machine = await db.machines.find_one({"id": machine_id}, {"_id": 0})
    n_meas = await db.measurements.count_documents({"machine_id": machine_id})
    await db.machines.delete_one({"id": machine_id})
    await db.measurements.delete_many({"machine_id": machine_id})
    if machine:
        await log_deletion(user, "máquina", f"TAG {machine.get('tag','')}", {"id": machine_id, "medicoes_removidas": n_meas})
    return {"ok": True}

@api.post("/machines/import")
async def import_machines(file: UploadFile = File(...), user=Depends(require_editor)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    inserted = 0
    skipped = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tipo = "vibracao" if "VIBRA" in sheet_name.upper() else ("termografia" if "TERMO" in sheet_name.upper() else "ambos")
        header_idx = {}
        headers_found = False
        # Forward-fill state
        cur_tag = None
        cur_local = None
        cur_equip = None
        for row in ws.iter_rows(values_only=True):
            if not headers_found:
                vals = [str(c).strip().upper() if c else "" for c in row]
                if "TAG" in vals:
                    for i, v in enumerate(vals):
                        header_idx[v] = i
                    headers_found = True
                continue
            tag_i = header_idx.get("TAG")
            local_i = header_idx.get("LOCAL DE INSTALAÇÃO", -1)
            equip_i = header_idx.get("EQUIPAMENTOS", -1)
            desc_i = header_idx.get("DESCRIÇÃO", -1)
            comp_i = header_idx.get("COMPONENTE", -1)
            rpm_i = header_idx.get("ROTAÇÃO", -1)
            pot_i = header_idx.get("POTENCIA", -1)
            rla_i = header_idx.get("ROLAMENTO LA", -1)
            rloa_i = header_idx.get("ROLAMENTO LOA", -1)

            raw_tag = row[tag_i] if tag_i is not None and tag_i < len(row) else None
            raw_local = row[local_i] if local_i >= 0 and local_i < len(row) else None
            raw_equip = row[equip_i] if equip_i >= 0 and equip_i < len(row) else None
            raw_desc = row[desc_i] if desc_i >= 0 and desc_i < len(row) else None
            raw_comp = row[comp_i] if comp_i >= 0 and comp_i < len(row) else None

            # Section header rows (only TAG cell, no local) → just remember TAG context but skip
            if raw_tag and not raw_local and not raw_desc and not raw_comp:
                # section header like "INDUSTRIA"; skip
                cur_tag, cur_local, cur_equip = None, None, None
                continue

            # New parent row?
            if raw_tag and raw_local:
                cur_tag = str(raw_tag).strip()
                cur_local = str(raw_local).strip()
                cur_equip = str(raw_equip or "").strip()

            if not cur_tag:
                continue

            is_primary = bool(raw_tag and raw_local)

            if tipo == "termografia":
                # Termografia: cada linha (TAG+LOCAL) é um ponto/equipamento próprio.
                # Componente é opcional; se vazio, a própria TAG é o registro.
                sub = str(raw_comp).strip() if raw_comp else ""
                if not is_primary and not sub:
                    continue
                full_tag = cur_tag if not sub else f"{cur_tag} / {sub}"
            else:
                # Vibração: subconjunto vem da Descrição (obrigatório)
                sub = (str(raw_desc).strip() if raw_desc else "") or (str(raw_comp).strip() if raw_comp else "")
                if not sub:
                    # No subconjunto info — skip (avoids duplicate parent rows)
                    continue
                full_tag = f"{cur_tag} / {sub}"

            try:
                doc = {
                    "id": str(uuid.uuid4()),
                    "tag": full_tag,
                    "parent_tag": cur_tag,
                    "subconjunto": sub,
                    "local": cur_local or "",
                    "equipamento": cur_equip or "",
                    "descricao": sub if tipo == "vibracao" else "",
                    "componente": sub if tipo == "termografia" else "",
                    "rpm": float(row[rpm_i]) if rpm_i >= 0 and rpm_i < len(row) and row[rpm_i] not in (None, "") else None,
                    "potencia": float(row[pot_i]) if pot_i >= 0 and pot_i < len(row) and row[pot_i] not in (None, "") else None,
                    "rolamento_loa": str(row[rloa_i] or "").strip() if rloa_i >= 0 and rloa_i < len(row) else "",
                    "rolamento_la": str(row[rla_i] or "").strip() if rla_i >= 0 and rla_i < len(row) else "",
                    "tipo": tipo,
                    "fabricante": "",
                    "criticidade": "Média",
                    "status": "Sem diag.",
                    "created_at": now_iso(),
                }
                exists = await db.machines.find_one({"tag": full_tag})
                if exists:
                    skipped += 1
                else:
                    await db.machines.insert_one(doc)
                    inserted += 1
            except Exception as ex:
                logger.warning("Skip row: %s", ex)
                skipped += 1
    return {"inserted": inserted, "skipped": skipped}

# ============== Defects ==============
@api.get("/defects")
async def list_defects(user=Depends(get_current_user)):
    items = await db.defects.find({}, {"_id": 0}).to_list(2000)
    return items

@api.post("/defects")
async def create_defect(payload: DefectModel, user=Depends(require_editor)):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = now_iso()
    await db.defects.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/defects/{defect_id}")
async def update_defect(defect_id: str, payload: DefectModel, user=Depends(require_editor)):
    update = {k: v for k, v in payload.model_dump().items() if k != "id"}
    res = await db.defects.update_one({"id": defect_id}, {"$set": update})
    if not res.matched_count:
        raise HTTPException(404, "Defeito não encontrado")
    return {"ok": True}

@api.delete("/defects/{defect_id}")
async def delete_defect(defect_id: str, user=Depends(require_editor)):
    d = await db.defects.find_one({"id": defect_id}, {"_id": 0})
    await db.defects.delete_one({"id": defect_id})
    if d:
        await log_deletion(user, "defeito", d.get("nome", ""), {"id": defect_id})
    return {"ok": True}

# ============== Diagnostics ==============
@api.get("/diagnostics")
async def list_diagnostics(q: Optional[str] = None, user=Depends(get_current_user)):
    qry: Dict[str, Any] = {}
    if q:
        qry["$or"] = [
            {"machine_tag": {"$regex": q, "$options": "i"}},
            {"diagnostico": {"$regex": q, "$options": "i"}},
            {"recomendacao": {"$regex": q, "$options": "i"}},
        ]
    items = await db.diagnostics.find(qry, {"_id": 0}).sort("data", -1).to_list(2000)
    return items

@api.post("/diagnostics")
async def create_diagnostic(payload: DiagnosticoModel, user=Depends(require_editor)):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["data"] = now_iso()
    doc["tecnico"] = doc.get("tecnico") or user.get("name", "")
    await db.diagnostics.insert_one(doc)
    # update machine status
    await db.machines.update_one({"id": doc["machine_id"]}, {"$set": {"status": doc["status"]}})
    doc.pop("_id", None)
    return doc

@api.put("/diagnostics/{diag_id}")
async def update_diagnostic(diag_id: str, payload: DiagnosticoModel, user=Depends(require_editor)):
    update = {k: v for k, v in payload.model_dump().items() if k not in ("id",)}
    res = await db.diagnostics.update_one({"id": diag_id}, {"$set": update})
    if not res.matched_count:
        raise HTTPException(404, "Diagnóstico não encontrado")
    await db.machines.update_one({"id": update.get("machine_id")}, {"$set": {"status": update.get("status")}})
    return {"ok": True}

@api.delete("/diagnostics/{diag_id}")
async def delete_diagnostic(diag_id: str, user=Depends(require_editor)):
    d = await db.diagnostics.find_one({"id": diag_id}, {"_id": 0})
    await db.diagnostics.delete_one({"id": diag_id})
    if d:
        await log_deletion(user, "diagnóstico", f"{d.get('machine_tag','')}: {(d.get('diagnostico','') or '')[:80]}", {"id": diag_id, "status": d.get("status")})
    return {"ok": True}

# ============== Measurements (Vibração) ==============
@api.get("/measurements")
async def list_measurements(machine_id: Optional[str] = None, user=Depends(get_current_user)):
    qry: Dict[str, Any] = {}
    if machine_id:
        qry["machine_id"] = machine_id
    items = await db.measurements.find(qry, {"_id": 0}).sort([("ordem", 1), ("data", 1)]).to_list(20000)
    for it in items:
        it.setdefault("ordem", 0)
        it["alarme"] = iso_alarm(it.get("valor"), it.get("unidade", ""), it.get("deteccao", ""))
    return items

@api.get("/audit/deletions")
async def list_deletions(user=Depends(require_admin)):
    items = await db.deletion_logs.find({}, {"_id": 0}).sort("data", -1).to_list(3000)
    return items

# ============== Manual (admin) ==============
MANUAL_PATH = ROOT_DIR.parent / "MANUAL.md"

@api.get("/manual")
async def get_manual(user=Depends(require_admin)):
    text = MANUAL_PATH.read_text(encoding="utf-8") if MANUAL_PATH.exists() else "# Manual\nManual não disponível."
    return {"markdown": text}

@api.get("/manual/pdf")
async def get_manual_pdf(user=Depends(require_admin)):
    from starlette.responses import StreamingResponse
    import markdown as _md
    from xhtml2pdf import pisa
    text = MANUAL_PATH.read_text(encoding="utf-8") if MANUAL_PATH.exists() else "# Manual\nManual não disponível."
    body = _md.markdown(text, extensions=["tables", "fenced_code"])
    html = f"""<html><head><meta charset="utf-8"><style>
      @page {{ size: A4; margin: 1.8cm; }}
      body {{ font-family: Helvetica, Arial, sans-serif; font-size: 10.5pt; color: #1e293b; line-height: 1.5; }}
      h1 {{ font-size: 20pt; color: #0f172a; border-bottom: 2px solid #0f172a; padding-bottom: 4px; }}
      h2 {{ font-size: 14pt; color: #0f172a; margin-top: 18px; border-bottom: 1px solid #cbd5e1; padding-bottom: 3px; }}
      h3 {{ font-size: 11.5pt; color: #334155; margin-top: 12px; }}
      table {{ border-collapse: collapse; width: 100%; margin: 8px 0; }}
      th, td {{ border: 1px solid #cbd5e1; padding: 5px 7px; font-size: 9.5pt; text-align: left; }}
      th {{ background: #0f172a; color: #ffffff; }}
      code {{ background: #f1f5f9; padding: 1px 3px; }}
      ul {{ margin: 4px 0 4px 16px; }}
    </style></head><body>{body}</body></html>"""
    buf = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html), dest=buf, encoding="utf-8")
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=manual_fs_solucoes.pdf"})

@api.get("/measurements/template")
async def measurements_template(user=Depends(get_current_user)):
    """Gera um .xlsx pré-preenchido com todas as máquinas de vibração."""
    machines = await db.machines.find(
        {"tipo": {"$in": ["vibracao", "ambos"]}}, {"_id": 0}
    ).sort("tag", 1).to_list(10000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Overall"
    headers = ["Equipamento", "Subconjunto", "Ponto", "Unidade", "Detecção", "Valor"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
    for m in machines:
        ws.append([
            m.get("parent_tag") or m.get("tag", ""),
            m.get("subconjunto", ""),
            "", "mm/s", "Velocidade", "",
        ])
    widths = [18, 22, 14, 10, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_overall_vibracao.xlsx"},
    )

@api.post("/measurements/import")
async def import_measurements(machine_id: Optional[str] = Query(None), file: UploadFile = File(...), user=Depends(require_editor)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    inserted = 0
    skipped = 0
    seq = 0
    batch_ts = now_iso()
    header_idx: Dict[str, int] = {}
    headers_found = False

    def col(row, key, default=None):
        i = header_idx.get(key, -1)
        if 0 <= i < len(row):
            return row[i]
        return default

    for row in ws.iter_rows(values_only=True):
        if not headers_found:
            vals = [str(c).strip() if c else "" for c in row]
            norm = [v.lower() for v in vals]
            if "equipamento" in norm or "ponto" in norm:
                for i, v in enumerate(vals):
                    header_idx[v.strip().lower()] = i
                headers_found = True
            continue
        try:
            equip = col(row, "equipamento")
            if equip in (None, ""):
                continue
            equip = str(equip).strip()
            sub = str(col(row, "subconjunto", "") or "").strip()
            ponto = str(col(row, "ponto", "") or "").strip()
            valor_raw = col(row, "valor")
            if valor_raw in (None, ""):
                continue
            valor = float(valor_raw)
            unidade = str(col(row, "unidade", "") or "").strip()
            deteccao = str(col(row, "detecção", "") or col(row, "deteccao", "") or "").strip()

            # Match machine: parent_tag + subconjunto (case-insensitive)
            machine = None
            if sub:
                machine = await db.machines.find_one({
                    "parent_tag": {"$regex": f"^{re.escape(equip)}$", "$options": "i"},
                    "subconjunto": {"$regex": f"^{re.escape(sub)}$", "$options": "i"},
                })
            if not machine and sub:
                # fuzzy: subconjunto "contém" (ex: "MOTOR" casa "Motor 01")
                machine = await db.machines.find_one({
                    "parent_tag": {"$regex": f"^{re.escape(equip)}$", "$options": "i"},
                    "subconjunto": {"$regex": re.escape(sub), "$options": "i"},
                })
            if not machine:
                # try exact composite tag
                machine = await db.machines.find_one({"tag": {"$regex": f"^{re.escape(equip)}$", "$options": "i"}})
            if not machine and sub:
                machine = await db.machines.find_one({"tag": {"$regex": f"^{re.escape(equip)} / {re.escape(sub)}$", "$options": "i"}})
            if not machine:
                # try parent_tag only
                machine = await db.machines.find_one({"parent_tag": {"$regex": f"^{re.escape(equip)}$", "$options": "i"}})
            if not machine:
                # create stub with composite tag
                full_tag = f"{equip} / {sub}" if sub else equip
                machine = {
                    "id": str(uuid.uuid4()),
                    "tag": full_tag,
                    "parent_tag": equip,
                    "subconjunto": sub,
                    "local": "",
                    "equipamento": equip,
                    "descricao": sub,
                    "tipo": "vibracao",
                    "criticidade": "Média",
                    "status": "Sem diag.",
                    "created_at": now_iso(),
                }
                await db.machines.insert_one(machine)

            doc = {
                "id": str(uuid.uuid4()),
                "machine_id": machine["id"],
                "machine_tag": machine.get("tag", equip),
                "subconjunto": sub or machine.get("subconjunto", ""),
                "ponto": ponto,
                "valor": valor,
                "unidade": unidade,
                "deteccao": deteccao,
                "ordem": seq,
                "data": batch_ts,
            }
            await db.measurements.insert_one(doc)
            inserted += 1
            seq += 1
        except Exception as ex:
            logger.warning("Skip measurement: %s", ex)
            skipped += 1
    return {"inserted": inserted, "skipped": skipped}

@api.post("/measurements")
async def create_measurement(payload: MeasurementModel, user=Depends(require_editor)):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["data"] = now_iso()
    await db.measurements.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.delete("/measurements/{meas_id}")
async def delete_measurement(meas_id: str, user=Depends(require_editor)):
    m = await db.measurements.find_one({"id": meas_id}, {"_id": 0})
    res = await db.measurements.delete_one({"id": meas_id})
    if not res.deleted_count:
        raise HTTPException(404, "Medição não encontrada")
    if m:
        await log_deletion(user, "medição", f"{m.get('machine_tag','')} • {m.get('ponto','')} ({m.get('deteccao','')}) = {m.get('valor','')} {m.get('unidade','')}", {"id": meas_id})
    return {"ok": True}

@api.delete("/measurements/point/clear")
async def delete_measurement_point(machine_id: str = Query(...), ponto: str = Query(""), deteccao: str = Query(""), user=Depends(require_editor)):
    """Remove todo o histórico de um ponto (máquina + ponto + detecção)."""
    sample = await db.measurements.find_one({"machine_id": machine_id, "ponto": ponto, "deteccao": deteccao}, {"_id": 0})
    res = await db.measurements.delete_many({"machine_id": machine_id, "ponto": ponto, "deteccao": deteccao})
    if res.deleted_count and sample:
        await log_deletion(user, "medição (ponto)", f"{sample.get('machine_tag','')} • {ponto} ({deteccao}) — {res.deleted_count} registro(s)", {"machine_id": machine_id, "ponto": ponto, "deteccao": deteccao})
    return {"ok": True, "deleted": res.deleted_count}

@api.get("/measurements/export")
async def export_measurements(machine_id: Optional[str] = Query(None), user=Depends(get_current_user)):
    from starlette.responses import StreamingResponse
    qry: Dict[str, Any] = {}
    if machine_id:
        qry["machine_id"] = machine_id
    meas = await db.measurements.find(qry, {"_id": 0}).to_list(20000)
    rowMap: Dict[str, Any] = {}
    colMeta: Dict[str, str] = {}
    for m in meas:
        colKey = _fmt_dt(m.get("data", ""))
        raw = m.get("data", "")
        if colKey not in colMeta or raw < colMeta[colKey]:
            colMeta[colKey] = raw
        key = f"{m['machine_id']}||{m.get('ponto','')}||{m.get('deteccao','')}"
        if key not in rowMap:
            rowMap[key] = {"machine_tag": m.get("machine_tag", ""), "ponto": m.get("ponto", ""), "deteccao": m.get("deteccao", ""), "unidade": m.get("unidade", ""), "ordem": m.get("ordem", 0), "values": {}}
        rowMap[key]["values"][colKey] = m.get("valor")
    cols = sorted(colMeta.keys(), key=lambda c: colMeta[c])
    buf = _pivot_xlsx(list(rowMap.values()), cols, "Vibracao")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=tabela_dados_vibracao.xlsx"})

# ============== Termografia (Temperatura) ==============
@api.get("/thermal")
async def list_thermal(machine_id: Optional[str] = None, user=Depends(get_current_user)):
    qry: Dict[str, Any] = {}
    if machine_id:
        qry["machine_id"] = machine_id
    items = await db.thermal.find(qry, {"_id": 0}).sort([("ordem", 1), ("data", 1)]).to_list(20000)
    for it in items:
        it.setdefault("ordem", 0)
        it["alarme"] = temp_alarm(it.get("temperatura"))
        try:
            it["delta_t"] = round(float(it.get("temperatura")) - float(it.get("temp_ambiente")), 1) if it.get("temp_ambiente") not in (None, "") else None
        except (TypeError, ValueError):
            it["delta_t"] = None
    return items

@api.get("/thermal/template")
async def thermal_template(user=Depends(get_current_user)):
    from starlette.responses import StreamingResponse
    machines = await db.machines.find({"tipo": {"$in": ["termografia", "ambos"]}}, {"_id": 0}).sort("tag", 1).to_list(10000)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Termografia"
    headers = ["Equipamento", "Ponto", "Temperatura", "Temp_Ambiente"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
    for m in machines:
        ws.append([m.get("tag", ""), m.get("componente", "") or "", "", ""])
    for i, w in enumerate([22, 18, 14, 14], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=template_termografia.xlsx"})

@api.post("/thermal/import")
async def import_thermal(file: UploadFile = File(...), user=Depends(require_editor)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    inserted = 0; skipped = 0; seq = 0
    batch_ts = now_iso()
    header_idx: Dict[str, int] = {}; headers_found = False

    def col(row, key, default=None):
        i = header_idx.get(key, -1)
        return row[i] if 0 <= i < len(row) else default

    for row in ws.iter_rows(values_only=True):
        if not headers_found:
            vals = [str(c).strip() if c else "" for c in row]
            norm = [v.lower() for v in vals]
            if "equipamento" in norm or "temperatura" in norm:
                for i, v in enumerate(vals):
                    header_idx[v.strip().lower()] = i
                headers_found = True
            continue
        try:
            equip = col(row, "equipamento")
            if equip in (None, ""):
                continue
            equip = str(equip).strip()
            ponto = str(col(row, "ponto", "") or "").strip()
            temp_raw = col(row, "temperatura")
            if temp_raw in (None, ""):
                continue
            temperatura = float(temp_raw)
            amb_raw = col(row, "temp_ambiente")
            temp_ambiente = float(amb_raw) if amb_raw not in (None, "") else None

            machine = await db.machines.find_one({"tag": {"$regex": f"^{re.escape(equip)}$", "$options": "i"}})
            if not machine:
                machine = await db.machines.find_one({"parent_tag": {"$regex": f"^{re.escape(equip)}$", "$options": "i"}})
            if not machine:
                machine = {"id": str(uuid.uuid4()), "tag": equip, "parent_tag": equip, "subconjunto": "", "local": "", "equipamento": equip, "descricao": "", "componente": ponto, "tipo": "termografia", "criticidade": "Média", "status": "Sem diag.", "created_at": now_iso()}
                await db.machines.insert_one(machine)

            doc = {"id": str(uuid.uuid4()), "machine_id": machine["id"], "machine_tag": machine.get("tag", equip),
                   "ponto": ponto, "temperatura": temperatura, "temp_ambiente": temp_ambiente, "ordem": seq, "data": batch_ts}
            await db.thermal.insert_one(doc)
            inserted += 1; seq += 1
        except Exception as ex:
            logger.warning("Skip thermal: %s", ex)
            skipped += 1
    return {"inserted": inserted, "skipped": skipped}

@api.delete("/thermal/point/clear")
async def delete_thermal_point(machine_id: str = Query(...), ponto: str = Query(""), user=Depends(require_editor)):
    sample = await db.thermal.find_one({"machine_id": machine_id, "ponto": ponto}, {"_id": 0})
    res = await db.thermal.delete_many({"machine_id": machine_id, "ponto": ponto})
    if res.deleted_count and sample:
        await log_deletion(user, "temperatura (ponto)", f"{sample.get('machine_tag','')} • {ponto} — {res.deleted_count} registro(s)", {"machine_id": machine_id, "ponto": ponto})
    return {"ok": True, "deleted": res.deleted_count}

@api.delete("/thermal/{tid}")
async def delete_thermal(tid: str, user=Depends(require_editor)):
    t = await db.thermal.find_one({"id": tid}, {"_id": 0})
    res = await db.thermal.delete_one({"id": tid})
    if not res.deleted_count:
        raise HTTPException(404, "Registro não encontrado")
    if t:
        await log_deletion(user, "temperatura", f"{t.get('machine_tag','')} • {t.get('ponto','')} = {t.get('temperatura','')}°C", {"id": tid})
    return {"ok": True}

@api.get("/thermal/export")
async def export_thermal(machine_id: Optional[str] = Query(None), user=Depends(get_current_user)):
    from starlette.responses import StreamingResponse
    qry: Dict[str, Any] = {}
    if machine_id:
        qry["machine_id"] = machine_id
    items = await db.thermal.find(qry, {"_id": 0}).to_list(20000)
    rowMap: Dict[str, Any] = {}
    colMeta: Dict[str, str] = {}
    for m in items:
        colKey = _fmt_dt(m.get("data", ""))
        raw = m.get("data", "")
        if colKey not in colMeta or raw < colMeta[colKey]:
            colMeta[colKey] = raw
        key = f"{m['machine_id']}||{m.get('ponto','')}"
        if key not in rowMap:
            rowMap[key] = {"machine_tag": m.get("machine_tag", ""), "ponto": m.get("ponto", ""), "deteccao": "Temperatura", "unidade": "°C", "ordem": m.get("ordem", 0), "values": {}}
        rowMap[key]["values"][colKey] = m.get("temperatura")
    cols = sorted(colMeta.keys(), key=lambda c: colMeta[c])
    buf = _pivot_xlsx(list(rowMap.values()), cols, "Termografia")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=tabela_dados_termografia.xlsx"})

# ============== Dashboard ==============
@api.get("/dashboard")
async def dashboard(user=Depends(get_current_user)):
    machines = await db.machines.find({}, {"_id": 0}).to_list(10000)
    diagnostics = await db.diagnostics.find({}, {"_id": 0}).to_list(5000)
    total = len(machines)
    machines_with_diag = {d["machine_id"] for d in diagnostics}
    status_dist = {"OK": 0, "A1": 0, "A2": 0, "Parado": 0, "Sem diag.": 0}
    for m in machines:
        if m["id"] not in machines_with_diag:
            status_dist["Sem diag."] += 1
        else:
            s = m.get("status", "OK")
            if s in status_dist:
                status_dist[s] += 1
            else:
                status_dist["OK"] += 1

    # Health index (weighted) — diagnosticadas
    weight = {"OK": 100, "A1": 70, "A2": 40, "Parado": 0}
    diag_machines = [m for m in machines if m["id"] in machines_with_diag]
    if diag_machines:
        health = sum(weight.get(m.get("status", "OK"), 0) for m in diag_machines) / len(diag_machines)
    else:
        health = 0

    # Top defects (from diagnostics)
    defect_counter: Dict[str, int] = {}
    for d in diagnostics:
        for did in d.get("defect_ids", []):
            defect_counter[did] = defect_counter.get(did, 0) + 1
    defects = await db.defects.find({}, {"_id": 0}).to_list(2000)
    id_to_name = {d["id"]: d["nome"] for d in defects}
    top_defects = sorted(
        [{"nome": id_to_name.get(k, k), "count": v} for k, v in defect_counter.items()],
        key=lambda x: x["count"], reverse=True
    )[:10]

    # Monthly evolution (last 6 months)
    months: Dict[str, Dict[str, int]] = {}
    for d in diagnostics:
        try:
            dt = datetime.fromisoformat(d["data"].replace("Z", "+00:00"))
            key = dt.strftime("%Y-%m")
            months.setdefault(key, {"OK": 0, "A1": 0, "A2": 0, "Parado": 0})
            months[key][d.get("status", "A1")] = months[key].get(d.get("status", "A1"), 0) + 1
        except Exception:
            pass
    monthly = [{"month": k, **v} for k, v in sorted(months.items())][-6:]

    # Failure probability ranking
    crit_w = {"Alta": 3, "Média": 2, "Baixa": 1}
    status_w = {"OK": 0, "A1": 30, "A2": 70, "Parado": 100}
    ranking = []
    for m in machines:
        score = status_w.get(m.get("status", "OK"), 0) * crit_w.get(m.get("criticidade", "Média"), 2)
        ranking.append({"tag": m.get("tag"), "equipamento": m.get("equipamento", ""), "local": m.get("local", ""), "status": m.get("status"), "criticidade": m.get("criticidade"), "score": score})
    ranking = sorted(ranking, key=lambda x: x["score"], reverse=True)[:10]

    # Alertas do último upload (medições/temperaturas em A2 ou Parado — último valor por ponto)
    alerts = []
    meas = await db.measurements.find({}, {"_id": 0}).to_list(20000)
    latest_m: Dict[str, Any] = {}
    for m in meas:
        key = f"{m['machine_id']}||{m.get('ponto','')}||{m.get('deteccao','')}"
        cur = latest_m.get(key)
        if not cur or (m.get("data", "") > cur.get("data", "")):
            latest_m[key] = m
    for m in latest_m.values():
        al = iso_alarm(m.get("valor"), m.get("unidade", ""), m.get("deteccao", ""))
        if al in ("A2", "Parado"):
            alerts.append({"tipo": "Vibração", "machine_tag": m.get("machine_tag", ""), "ponto": m.get("ponto", ""),
                           "valor": m.get("valor"), "unidade": m.get("unidade", ""), "alarme": al, "data": m.get("data")})
    thermal = await db.thermal.find({}, {"_id": 0}).to_list(20000)
    latest_t: Dict[str, Any] = {}
    for t in thermal:
        key = f"{t['machine_id']}||{t.get('ponto','')}"
        cur = latest_t.get(key)
        if not cur or (t.get("data", "") > cur.get("data", "")):
            latest_t[key] = t
    for t in latest_t.values():
        al = temp_alarm(t.get("temperatura"))
        if al in ("A2", "Parado"):
            alerts.append({"tipo": "Termografia", "machine_tag": t.get("machine_tag", ""), "ponto": t.get("ponto", ""),
                           "valor": t.get("temperatura"), "unidade": "°C", "alarme": al, "data": t.get("data")})
    alerts = sorted(alerts, key=lambda x: (x["alarme"] != "Parado", x.get("data") or ""), reverse=True)[:15]

    return {
        "total_machines": total,
        "status_dist": status_dist,
        "health_index": round(health, 1),
        "top_defects": top_defects,
        "monthly_evolution": monthly,
        "failure_ranking": ranking,
        "latest_alerts": alerts,
        "total_diagnostics": len(diagnostics),
    }

# ============== Reports ==============
@api.get("/reports/summary")
async def report_summary(user=Depends(get_current_user)):
    """Resumo executivo: KPIs + status + top defeitos."""
    machines = await db.machines.find({}, {"_id": 0}).to_list(10000)
    diagnostics = await db.diagnostics.find({}, {"_id": 0}).sort("data", -1).to_list(5000)
    defects = await db.defects.find({}, {"_id": 0}).to_list(2000)
    id_to_name = {d["id"]: d["nome"] for d in defects}
    status_dist = {"OK": 0, "A1": 0, "A2": 0, "Parado": 0}
    for m in machines:
        status_dist[m.get("status", "OK")] = status_dist.get(m.get("status", "OK"), 0) + 1
    weight = {"OK": 100, "A1": 70, "A2": 40, "Parado": 0}
    health = sum(weight.get(m.get("status", "OK"), 0) for m in machines) / max(len(machines), 1)
    defect_counter: Dict[str, int] = {}
    for d in diagnostics:
        for did in d.get("defect_ids", []):
            defect_counter[did] = defect_counter.get(did, 0) + 1
    top_defects = sorted(
        [{"nome": id_to_name.get(k, k), "count": v} for k, v in defect_counter.items()],
        key=lambda x: x["count"], reverse=True,
    )[:5]
    critical = [
        {"tag": m.get("tag"), "equipamento": m.get("equipamento", ""), "local": m.get("local", ""), "status": m.get("status")}
        for m in machines if m.get("status") in ("A2", "Parado")
    ]
    return {
        "generated_at": now_iso(),
        "total_machines": len(machines),
        "status_dist": status_dist,
        "health_index": round(health, 1),
        "top_defects": top_defects,
        "critical_machines": critical,
        "total_diagnostics": len(diagnostics),
    }

@api.get("/reports/complete")
async def report_complete(user=Depends(get_current_user)):
    """Relatório completo: tudo do resumo + lista de todos os diagnósticos com detalhes."""
    summary = await report_summary(user)
    diagnostics = await db.diagnostics.find({}, {"_id": 0}).sort("data", -1).to_list(5000)
    machines = await db.machines.find({}, {"_id": 0}).to_list(10000)
    machine_map = {m["id"]: m for m in machines}
    enriched = []
    for d in diagnostics:
        m = machine_map.get(d.get("machine_id"), {})
        enriched.append({**d, "equipamento": m.get("equipamento", ""), "local": m.get("local", ""), "criticidade": m.get("criticidade", "")})
    return {**summary, "diagnostics": enriched, "machines": machines}

@api.get("/machines/{machine_id}/history")
async def machine_history(machine_id: str, user=Depends(get_current_user)):
    diags = await db.diagnostics.find({"machine_id": machine_id}, {"_id": 0}).sort("data", 1).to_list(1000)
    return diags


# ============== Mount ==============
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
